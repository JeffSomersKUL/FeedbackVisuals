# import openpyxl module
from csv import reader
import openpyxl
import csv


def get_column_number(col_letter):
    return ord(col_letter) - 64


def represents_int(s):
    try:
        int(s)
        return True
    except ValueError:
        return False


#########################################################
# Waarden van de thema's ################################
plan_van_aanpak = 0
concepten = 1
wiskundig_model = 2
rekentechnische_aanpak = 3
fysische_interpretatie = 4
#########################################################

#########################################################
# known data from template ##############################
column_questions_letter = 'A'
column_questions_number = get_column_number(column_questions_letter)
column_thema_letter = 'B'
column_thema_number = get_column_number(column_thema_letter)
column_kind_of_question_letter = 'C'
column_kind_of_question_number = get_column_number(column_kind_of_question_letter)
column_weight_question_letter = 'D'
column_weight_question_number = get_column_number(column_weight_question_letter)
column_maximal_score_letter = 'E'
column_maximale_score_number = get_column_number(column_maximal_score_letter)
#########################################################

#########################################################
# known data from survey ##############################
column_response_ID_letter = 'I'
column_response_ID_number = get_column_number(column_response_ID_letter)
row_question_number = 1
#########################################################

###################
multiple = 0
one_possible = 1
drag = 2
###################

# return the row number as in excel
def get_row_student(response_id, sheet_servey, column_len_survey):
    row_index = 3
    while row_index <= column_len_survey:
        value_row_id = sheet_servey.cell(row=row_index, column = column_response_ID_number).value
        if value_row_id == response_id:
            return row_index
        row_index += 1
    return None


def get_normal_question(question, sheet_servey, row_len_survey):
    column_index = 14
    while column_index <= row_len_survey:
        value_column_index = sheet_servey.cell(row=row_question_number, column=column_index).value
        if value_column_index == question:
            return column_index
        column_index += 1
    return None


def get_column_drag_question(question, sheet_servey, row_len_survey):
    column_index = 14
    question = str(question + '_GROUP')
    while column_index <= row_len_survey:
        value_column_index = sheet_servey.cell(row=row_question_number, column=column_index).value
        if value_column_index == question:
            return column_index
        column_index += 1
    return None

# return the column number as in excel
def get_column_question(question, sheet_servey, row_len_survey, kind):
    if kind == drag:
        return get_column_drag_question(question, sheet_servey, row_len_survey)
    else:
        return get_normal_question(question, sheet_servey, row_len_survey)

def get_total_score(list_of_score):
    full_score = 0
    for score in list_of_score:
        full_score += score
    return full_score / len(list_of_score)


def calculate_score_multiple(index_question, number_question_offsets, maximale_score, sheet_template):
    sum_score = 0
    for offset in number_question_offsets:
        if represents_int(offset):
            offset_score = sheet_template.cell(row=index_question, column=int(column_maximale_score_number) +
                                                                          int(offset)).value
            sum_score += offset_score
    var_score = float(sum_score) / float(maximale_score)
    if var_score > 1:
        return 1
    else:
        return var_score


def calculate_score_single(index_question, number_question_offset, maximale_score, sheet_template):
    weighted_score = sheet_template.cell(row=index_question, column=int(column_maximale_score_number) +
                                                                    int(number_question_offset)).value
    return float(weighted_score) / float(maximale_score)


def change_to_nr_list(list_with_strings):
    if list_with_strings:
        for i in list_with_strings:
            i = int(i)


def calculate_score_drag(index_question, answer, sheet_template):
    correct_answer = sheet_template.cell(row=index_question, column=int(column_maximale_score_number) + int(1)).value
    score = 0
    answer = str(answer).split(",")
    change_to_nr_list(answer)
    correct_answer = str(correct_answer).split(".")
    change_to_nr_list(correct_answer)

    if len(correct_answer) > 1:
        if set(answer) == set(correct_answer):
            score = 1
    elif len(answer) > 1:
        if correct_answer[0] in answer:
            score = 1
    else:
        if correct_answer == answer:
            score = 1

    return score




def get_score(kind, index_question, answer, maximale_score, sheet_template):
    if answer != None and answer != '':
        score = 0
        if kind == one_possible:
            score = calculate_score_single(index_question, answer, maximale_score, sheet_template)
        elif kind == multiple:
            score = calculate_score_multiple(index_question, answer, maximale_score, sheet_template)
        elif kind == drag:
            score = calculate_score_drag(index_question, answer, sheet_template)
        return score
    else:
        return 0

# gets the data from the excel file from one student
def get_list_data(student, sheet_template, sheet_servey, column_len_template, column_len_survey, row_len_survey, sum_weights):
    data_plan_van_aanpak = 0
    data_concepten = 0
    data_wiskundig_model = 0
    data_rekentechnische_aanpak = 0
    data_fysische_interpretatie = 0

    row_index_questions = 2
    row_student = get_row_student(student, sheet_servey, column_len_survey)  # kan nog sneller door mee de rij van de student al doortegeven ipv die hier terug opnieuw op te zoeken
    # for each question in the template we search the correct score for the student and add it to the correct thema
    while row_index_questions <= column_len_template:
        question = sheet_template.cell(row=row_index_questions, column=column_questions_number).value  # template
        thema = sheet_template.cell(row=row_index_questions, column=column_thema_number).value  # template
        kind = sheet_template.cell(row=row_index_questions, column=column_kind_of_question_number).value  # template
        maximale_score = sheet_template.cell(row=row_index_questions, column=column_maximale_score_number).value  # template
        weight = sheet_template.cell(row=row_index_questions, column=column_weight_question_number).value  # template

        question_column_number = get_column_question(question, sheet_servey, row_len_survey, kind)  # survey
        answer_student = sheet_servey.cell(row=row_student, column=question_column_number).value  # survey


        score = get_score(kind, row_index_questions, answer_student, maximale_score, sheet_template)  # template

        # add it to the right thema
        if int(thema) == plan_van_aanpak:
            data_plan_van_aanpak += (weight/sum_weights[thema])*score
        elif int(thema) == concepten:
            data_concepten += (weight/sum_weights[thema])*score
        elif int(thema) == wiskundig_model:
            data_wiskundig_model += (weight/sum_weights[thema])*score
        elif int(thema) == rekentechnische_aanpak:
            data_rekentechnische_aanpak += (weight/sum_weights[thema])*score
        elif int(thema) == fysische_interpretatie:
            data_fysische_interpretatie += (weight/sum_weights[thema])*score
        row_index_questions += 1

    data = [data_plan_van_aanpak, data_concepten, data_wiskundig_model, data_rekentechnische_aanpak, data_fysische_interpretatie]
    return data

def get_total_weights_themas(sheet_template, column_len_template):
    weight_sum_plan_van_aanpak = 0
    weight_sum_concepten = 0
    weight_sum_wiskundig_model = 0
    weight_sum_rekentechnische_aanpak = 0
    weight_sum_fysische_interpretatie = 0

    row_index_questions = 2

    # for each question in the template we search the weight of the question and add them up
    while row_index_questions <= column_len_template:
        thema = sheet_template.cell(row=row_index_questions, column=column_thema_number).value  # template
        weight = sheet_template.cell(row=row_index_questions, column=column_weight_question_number).value  # template

        # add it to the right thema
        if int(thema) == plan_van_aanpak:
            weight_sum_plan_van_aanpak += weight
        elif int(thema) == concepten:
            weight_sum_concepten += weight
        elif int(thema) == wiskundig_model:
            weight_sum_wiskundig_model += weight
        elif int(thema) == rekentechnische_aanpak:
            weight_sum_rekentechnische_aanpak += weight
        elif int(thema) == fysische_interpretatie:
            weight_sum_fysische_interpretatie += weight
        row_index_questions += 1


    sum_weights = [weight_sum_plan_van_aanpak, weight_sum_concepten, weight_sum_wiskundig_model, weight_sum_rekentechnische_aanpak, weight_sum_fysische_interpretatie]
    return sum_weights


#########################################################
# known data from survey ##############################
length_r_number = 7
column_first_letter_rnumber_letter = 'R'
column_first_letter_rnumber_number = get_column_number(column_first_letter_rnumber_letter)
column_first_letter_rnumber_not_answered_letter = 'Y'
column_first_letter_rnumber_not_answered_number = get_column_number(column_first_letter_rnumber_not_answered_letter)
#########################################################


def get_rnummer_of_row(row, sheet_servey):
    rnumber = 'r'
    index = 0
    while index < length_r_number:
        number = sheet_servey.cell(row=row, column=column_first_letter_rnumber_number+index).value
        if number == '':
            rnumber_2 = 'r'
            index_2 = 0
            while index_2 < length_r_number:
                number_2 = sheet_servey.cell(row=row, column=column_first_letter_rnumber_not_answered_number+index).value
                if number_2 == '':
                    return 'r0000000'
                number_2 = int(number_2)
                number_2 = number_2 - 1
                rnumber_2 += str(number_2)
                index_2 += 1
            return rnumber_2
        number = int(number)
        number = number - 1  # for some reason Qualtrics gives the number you give in one extra value
        rnumber += str(int(number))
        index += 1
    return rnumber

def save_rnummeer_student(rnummer):
    already_saved = False
    with open('student.csv', 'r') as read_obj:
        csv_reader = reader(read_obj)
        for row in csv_reader:
            if row[0] == rnummer:
                already_saved = True
                break

    if not already_saved:
        with open('student.csv', 'a', newline='') as file:
            writer = csv.writer(file)
            writer.writerow([rnummer])


def create_csv_file(servey, template, name):
    servey_file = 'werkzittingen/' + servey
    template_file = 'werkzittingen/' + template


    # To open the workbook
    # workbook object is created
    wb_servey = openpyxl.load_workbook(servey_file)
    wb_template = openpyxl.load_workbook(template_file)

    # Get workbook active sheet object
    # from the active attribute
    sheet_servey = wb_servey.active
    sheet_template = wb_template.active

    # Cell objects also have a row, column,
    # and coordinate attributes that provide
    # location information for the cell.

    # Note: The first row or
    # column integer is 1, not 0.

    # Cell object is created by using
    # sheet object's cell() method.

    column_len_template = len(list(sheet_template.columns)[0])

    column_len_survey = len(list(sheet_servey.columns)[0])
    row_len_survey = len(list(sheet_servey.rows)[0])

    sum_weights = get_total_weights_themas(sheet_template, column_len_template)

    sum_plan_van_aanpak = 0
    sum_concepten = 0
    sum_wiskundig_model = 0
    sum_rekentechnische_aanpak = 0
    sum_fysische_interpretatie = 0
    sum_student = 0

    with open('../feedback-server/oefenzittingen/' + name + '.csv', 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Rnummer"]+["plan"]+["concepten"]+["wiskundig"]+["rekentechnisch"]+["interpretatie"])

        row_index = 3
        while row_index < column_len_survey:
            student = sheet_servey.cell(row=row_index, column=column_response_ID_number).value
            rnummer_student = get_rnummer_of_row(row_index, sheet_servey)
            save_rnummeer_student(rnummer_student)
            data_student = get_list_data(student, sheet_template, sheet_servey, column_len_template, column_len_survey, row_len_survey, sum_weights)
            writer.writerow([rnummer_student] + data_student)
            sum_plan_van_aanpak += data_student[plan_van_aanpak]
            sum_concepten += data_student[concepten]
            sum_wiskundig_model += data_student[wiskundig_model]
            sum_rekentechnische_aanpak += data_student[rekentechnische_aanpak]
            sum_fysische_interpretatie += data_student[fysische_interpretatie]
            sum_student += 1
            row_index += 1


        average_data = [float(sum_plan_van_aanpak)/float(sum_student), float(sum_concepten)/float(sum_student),
                        float(sum_wiskundig_model)/float(sum_student), float(sum_rekentechnische_aanpak)/float(sum_student),
                        float(sum_fysische_interpretatie)/float(sum_student)]

        writer.writerow(['Average'] + average_data)

create_csv_file('data_wz8.xlsx','template_wz8.xlsx', 'wz8')